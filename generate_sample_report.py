import pandas as pd
import random
from datetime import datetime
import xlsxwriter
import sys
import os

def create_report():
    file_name = "report.xlsx"
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
    workbook = writer.book

    # Define Formats
    header_format = workbook.add_format({
        'bold': True, 'text_wrap': True, 'valign': 'top',
        'fg_color': '#D7E4BC', 'border': 1
    })
    
    pass_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
    fail_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
    skip_format = workbook.add_format({'bg_color': '#FFEB9C', 'font_color': '#9C6500'})
    
    crit_format = workbook.add_format({'bg_color': '#8B0000', 'font_color': '#FFFFFF'})
    high_format = workbook.add_format({'bg_color': '#FF0000', 'font_color': '#FFFFFF'})
    med_format = workbook.add_format({'bg_color': '#FFA500', 'font_color': '#FFFFFF'})
    low_format = workbook.add_format({'bg_color': '#FFFF00', 'font_color': '#000000'})
    info_format = workbook.add_format({'bg_color': '#ADD8E6', 'font_color': '#000000'})
    
    cell_format = workbook.add_format({'border': 1, 'text_wrap': True, 'valign': 'top'})
    
    # ------------------
    # Data Generation
    # ------------------
    # Mobile Tests
    mobile_data = []
    for i in range(1, 301):
        mobile_data.append({
            'Test ID': f'TC-MOB-{i:03d}',
            'Test Name': f'Mobile Test {i}',
            'Module': random.choice(['Login', 'Checkout', 'Profile']),
            'Platform': random.choice(['Android', 'iOS']),
            'Device Name': random.choice(['Pixel 7', 'iPhone 14', 'Galaxy S23']),
            'Device Model': 'Generic',
            'OS Version': random.choice(['13.0', '16.4', '14.0']),
            'App Version': '1.0.0',
            'Status': 'Pass',
            'Execution Time': f'{random.uniform(0.5, 3.0):.2f}s',
            'Screenshot': 'N/A',
            'Error Message': '',
            'Remarks': 'Passed successfully'
        })
    df_mobile = pd.DataFrame(mobile_data)
    
    # Web Tests
    web_data = []
    for i in range(1, 301):
        web_data.append({
            'Test ID': f'TC-WEB-{i:03d}',
            'Test Name': f'Web Test {i}',
            'Module': random.choice(['Search', 'Cart', 'Payment']),
            'Browser': random.choice(['Chrome', 'Firefox', 'Safari']),
            'Browser Version': 'Latest',
            'URL': 'https://example.com/test',
            'Status': 'Pass',
            'Execution Time': f'{random.uniform(0.2, 2.0):.2f}s',
            'Screenshot': 'N/A',
            'Error Message': '',
            'Remarks': 'Passed successfully'
        })
    df_web = pd.DataFrame(web_data)
    
    # Load Tests
    load_data = []
    for i in range(1, 301):
        load_data.append({
            'Scenario': f'Scenario {i}',
            'Virtual Users': random.randint(50, 500),
            'Ramp-Up Time': f'{random.randint(10, 60)}s',
            'Average Response Time': f'{random.randint(100, 500)}ms',
            'Maximum Response Time': f'{random.randint(600, 2000)}ms',
            'Minimum Response Time': f'{random.randint(20, 90)}ms',
            'Throughput': f'{random.randint(50, 200)} req/s',
            'Requests Per Second': random.randint(50, 200),
            'Error Percentage': '0%',
            'CPU Usage': f'{random.randint(20, 60)}%',
            'Memory Usage': f'{random.randint(30, 70)}%',
            'Status': 'Pass'
        })
    df_load = pd.DataFrame(load_data)
    
    # Vulnerability Tests
    vuln_data = []
    for i in range(1, 301):
        vuln_data.append({
            'Scan ID': f'SCAN-{i:03d}',
            'Module': random.choice(['API', 'Web', 'Database']),
            'Target': 'https://example.com/api',
            'Vulnerability': random.choice([
                'SQL Injection', 'Cross-Site Scripting (XSS)', 'CSRF',
                'Command Injection', 'Broken Authentication',
                'Missing Security Headers', 'Sensitive Data Exposure', 'Open Redirect'
            ]),
            'Severity': random.choice(['Critical', 'High', 'Medium', 'Low', 'Informational']),
            'CVE': 'N/A',
            'CVSS Score': f'{random.uniform(1.0, 10.0):.1f}',
            'OWASP Category': 'A6:2021',
            'Status': 'Pass',
            'Recommendation': 'Review config',
            'Scan Date': datetime.now().strftime('%Y-%m-%d')
        })
    df_vuln = pd.DataFrame(vuln_data)
    
    # Summary Data
    total_mobile = len(df_mobile)
    total_web = len(df_web)
    total_load = len(df_load)
    total_vuln = len(df_vuln)
    total_tests = total_mobile + total_web + total_load + total_vuln
    
    passed = total_tests
    failed = 0
    skipped = 0
    
    summary_metrics = {
        'Metric': [
            'Total Test Cases', 'Passed', 'Failed', 'Skipped', 
            'Pass Rate', 'Fail Rate', 'Execution Time', 
            'Environment', 'Build Number', 'Report Version', 'Date Generated'
        ],
        'Value': [
            total_tests, passed, failed, skipped,
            '100%', '0%', '15m 30s',
            'Staging', 'v1.2.34', '1.0', datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
    }
    df_summary = pd.DataFrame(summary_metrics)
    
    summary_table = {
        'Test Type': ['Mobile Tests', 'Web Tests', 'Load Tests', 'Vulnerability Tests'],
        'Total': [total_mobile, total_web, total_load, total_vuln],
        'Passed': [total_mobile, total_web, total_load, total_vuln],
        'Failed': [0, 0, 0, 0],
        'Skipped': [0, 0, 0, 0]
    }
    df_summary_table = pd.DataFrame(summary_table)

    # ------------------
    # Writing to Excel
    # ------------------
    
    # Write empty DataFrames just to get sheet structure if we want, but writing DataFrames is easier
    df_summary.to_excel(writer, sheet_name='Summary', index=False, startrow=1)
    df_summary_table.to_excel(writer, sheet_name='Summary', index=False, startrow=15)
    
    df_mobile.to_excel(writer, sheet_name='Mobile Tests', index=False)
    df_web.to_excel(writer, sheet_name='Web Tests', index=False)
    df_load.to_excel(writer, sheet_name='Load Tests', index=False)
    df_vuln.to_excel(writer, sheet_name='Vulnerability Tests', index=False)
    
    # ------------------
    # Formatting function
    # ------------------
    def format_sheet(sheet_name, df, start_row=0):
        worksheet = writer.sheets[sheet_name]
        
        # Add table
        cols = [{'header': c} for c in df.columns]
        worksheet.add_table(start_row, 0, start_row + len(df), len(df.columns) - 1, {
            'columns': cols,
            'style': 'Table Style Medium 9',
            'autofilter': True
        })
        
        # Auto-fit columns and freeze top row
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.set_column(idx, idx, max_len, cell_format)
            
        worksheet.freeze_panes(start_row+1, 0)
        
        # Conditional formatting for Status
        if 'Status' in df.columns:
            status_col = df.columns.get_loc('Status')
            col_letter = chr(65 + status_col)
            cell_range = f'{col_letter}{start_row+2}:{col_letter}{start_row+len(df)+1}'
            worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '==', 'value': '"Pass"', 'format': pass_format})
            worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '==', 'value': '"Fail"', 'format': fail_format})
            worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '==', 'value': '"Skipped"', 'format': skip_format})
            
        if 'Severity' in df.columns:
            sev_col = df.columns.get_loc('Severity')
            col_letter = chr(65 + sev_col)
            cell_range = f'{col_letter}{start_row+2}:{col_letter}{start_row+len(df)+1}'
            worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '==', 'value': '"Critical"', 'format': crit_format})
            worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '==', 'value': '"High"', 'format': high_format})
            worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '==', 'value': '"Medium"', 'format': med_format})
            worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '==', 'value': '"Low"', 'format': low_format})
            worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '==', 'value': '"Informational"', 'format': info_format})

    
    format_sheet('Mobile Tests', df_mobile)
    format_sheet('Web Tests', df_web)
    format_sheet('Load Tests', df_load)
    format_sheet('Vulnerability Tests', df_vuln)
    
    # Summary sheet formatting
    worksheet_sum = writer.sheets['Summary']
    worksheet_sum.set_column('A:A', 25, cell_format)
    worksheet_sum.set_column('B:B', 30, cell_format)
    
    # Add a Chart to Summary
    chart = workbook.add_chart({'type': 'column'})
    chart.add_series({
        'name': 'Passed',
        'categories': ['Summary', 16, 0, 19, 0],
        'values':     ['Summary', 16, 2, 19, 2],
    })
    chart.set_title({'name': 'Test Breakdown'})
    worksheet_sum.insert_chart('E4', chart)

    writer.close()

def verify_workbook():
    file_name = "report.xlsx"
    if not os.path.exists(file_name):
        print("Validation Failed: report.xlsx does not exist.")
        sys.exit(1)
        
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_name)
    except Exception as e:
        print(f"Validation Failed: Workbook cannot be opened. {e}")
        sys.exit(1)
        
    sheets = wb.sheetnames
    expected = ['Summary', 'Mobile Tests', 'Web Tests', 'Load Tests', 'Vulnerability Tests']
    if sheets != expected:
        print(f"Validation Failed: Worksheets are not in expected order or missing. Found: {sheets}")
        sys.exit(1)
        
    for s in expected:
        if wb[s].max_row <= 1 and s != 'Summary':
            print(f"Validation Failed: Sheet {s} is empty.")
            sys.exit(1)
            
    print("Validation passed. Workbook created successfully.")
    sys.exit(0)

if __name__ == '__main__':
    create_report()
    verify_workbook()
